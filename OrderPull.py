import openpyxl
import os
import shutil
from sys import exit

# Check if Input.txt exists, if not, create it
input_file_path = "Input.txt"
if not os.path.exists(input_file_path):
    open(input_file_path, "w", encoding="utf-8")
    print("Please put input into the Input.txt file")

    # Wait for user to press Enter
    input("Press Enter to exit...")
    exit()
else:
    with open(input_file_path, 'r') as file:
        content = file.read()
        if len(content) == 0:
            print("The Input.txt file is empty")
            print("Please put input into the Input.txt file")

            # Wait for user to press Enter
            input("Press Enter to exit...")
            exit()


# Function to clean up data
def clean_up_data(data):
    cleaned_data = []
    for line in data:
        cleaned_line = line.replace("\t", "").replace(" ", "").replace("\n", "").lower()
        if cleaned_line:
            cleaned_data.append(cleaned_line)
    return cleaned_data

class PricingItem:
    def __init__(self, names, original_names, base_price, medium_price, big_price):
        self.names = names
        self.original_names = original_names
        self.base_price = base_price
        self.medium_price = medium_price
        self.big_price = big_price

class OrderItem:
    def __init__(self, menu_item, quantity, size, unit_price, additions):
        self.menu_item = menu_item
        self.size = size
        self.additions = additions
        self.unit_price = unit_price
        self.quantity = quantity

def load_data_from_excel(file_path, start_column):
    data = []

    # Load data from Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    num_columns = 4  # Number of columns for menu data (name, base price, medium price, big price)

    for row in sheet.iter_rows(min_row=2, min_col=start_column, max_col=start_column + num_columns - 1, values_only=True):
        if(row[0]) != None:
            raw_names = row[0].split(',')
            cleaned_names = clean_up_data(raw_names)
            original_names = raw_names
            base_price = row[1]
            medium_price = row[2]
            big_price = row[3]
            
            item = PricingItem(cleaned_names, original_names, base_price, medium_price, big_price)
            data.append(item)

    return data

def load_data_from_file(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        lines = file.readlines()

    cleaned_lines = clean_up_data(lines)
    order_as_text_list = [line for line in cleaned_lines if line.strip()]
    return order_as_text_list

# Provide the paths to internal files
internals_folder = "Internals"
excel_file_path = os.path.join(internals_folder, "Pricing.xlsx")
empty_order_file_path = os.path.join(internals_folder, "EmptyOrder.xlsx")

# Copy the EmptyOrder.xlsx to create a new Order.xlsx
if os.path.exists("Order.xlsx"):
    os.remove("Order.xlsx")
shutil.copy(empty_order_file_path, "Order.xlsx")

# Load main menu data from Excel
main_menu_start_column = 2  # Adjust this based on the desired starting column for main menu

menu = load_data_from_excel(excel_file_path, main_menu_start_column)

# Load additional menu data from Excel
additional_menu_start_column = 7  # Adjust this based on the desired starting column for additional menu
additions = load_data_from_excel(excel_file_path, additional_menu_start_column)

# Load data from Input.txt
order_as_text_list = load_data_from_file(input_file_path)

# Create OrderItem instances and add them to order_items list
order_items = []
for order_text in order_as_text_list:
    menu_item = None
    quantity = 1
    size = 0  # Defaults to base price
    unit_price = 0.0  # Default unit price

    # Determine the size from the order text
    if "mała" in order_text:
        size = 1  # Small
    if "średnia" in order_text:
        size = 2  # Medium
    elif "duża" in order_text:
        size = 3  # Large

    # Extract the quantity from the order text
    quantity_indicator = "x"
    if quantity_indicator in order_text:
        quantity_start = order_text.find(quantity_indicator)
        quantity_str = order_text[quantity_start + 1:]

        # Check if the following characters are a valid number
        if quantity_str.isdigit():
            quantity = int(quantity_str)
            order_text = order_text[:quantity_start]  # Remove the quantity indicator from order_text

    # Check if the order has additions
    order_additions = []
    if "+" in order_text:
        for addition in additions:
            cnt = 0

            for name in addition.names:
                cnt+= order_text.count(name)

            for num in range(0,cnt):
                order_additions.append(addition)

    for item in menu:
        if any(name in order_text for name in item.names):
            menu_item = item
            break

    if menu_item:
        # Determine the unit price based on the size
        additional_price = 0
        if len(order_additions):
            for addition in order_additions:
                if size == 2 and addition.medium_price != None:
                    additional_price += addition.medium_price
                elif size == 3 and addition.big_price != None:
                    additional_price += addition.big_price
                else:
                    additional_price += addition.base_price

        if size <= 1:
            unit_price += float(menu_item.base_price + additional_price)
        elif size == 2:
            unit_price += float(menu_item.medium_price + additional_price)
        elif size == 3:
            unit_price += float(menu_item.big_price + additional_price)

        order_item = OrderItem(menu_item, quantity, size, unit_price, order_additions)
        order_items.append(order_item)

# Open the existing Order.xlsx file
order_workbook = openpyxl.load_workbook("Order.xlsx")
order_sheet = order_workbook.active

# Write data from order_items list to the worksheet
for row_index, order_item in enumerate(order_items, start=2):
    order_sheet.cell(row=row_index, column=2, value=order_item.menu_item.original_names[0])

    # Determine the size text
    size_text = ""

    if order_item.size == 1:
        size_text = "mała"
    elif order_item.size == 2:
        size_text = "średnia"
    elif order_item.size == 3:
        size_text = "duża"
    order_sheet.cell(row=row_index, column=3, value=size_text)
    
    if len(order_item.additions):
        out_name=""
        for name_ in order_item.additions:
            out_name += "+" + ", ".join(name_.original_names)
        order_sheet.cell(row=row_index, column=4, value=out_name)

    order_sheet.cell(row=row_index, column=5, value=order_item.unit_price)
    order_sheet.cell(row=row_index, column=6, value=order_item.quantity)

# Save the changes to the existing Order.xlsx file
order_workbook.save("Order.xlsx")

# Copy Order.xlsx to PastOrders folder with the appropriate naming
past_orders_folder = "PastOrders"
existing_orders = [filename for filename in os.listdir(past_orders_folder) if filename.startswith("Order_")]
latest_order_number = 0

for filename in existing_orders:
    order_number = int(filename.split("_")[1].split(".")[0])
    if order_number > latest_order_number:
        latest_order_number = order_number

new_order_number = latest_order_number + 1
new_order_filename = f"Order_{new_order_number}.xlsx"
shutil.copy2("Order.xlsx", os.path.join(past_orders_folder, new_order_filename))

print("Order data has beed updated in Order.xlsx and has been put into the PastOrders folder")
input("Press Enter to exit...")